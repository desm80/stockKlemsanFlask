import os
from zipfile import ZipFile

import requests
import xlrd
from openpyxl.workbook import Workbook

from settings import BASE_DIR, URL_EPARH, URL_GLOBAL


def clear_folder(folder):
    for f in os.listdir(folder):
        os.remove(os.path.join(folder, f))


# def get_stock(part_number):
#     result = []
#     wb = load_workbook('./downloads/eparh.xlsx')
#     sheet = wb[wb.sheetnames[0]]
#     wb2 = load_workbook('./downloads/global.xlsx')
#     sheet2 = wb2[wb.sheetnames[0]]
#     for item in sheet.values:
#         if str(item[0]).find(str(part_number)) != -1:
#             result.append(
#                 f'Артикул {item[0]} в количестве {item[-1]} на складе ЭПАРХ')
#
#     for item in sheet2.values:
#         if str(item[0]).replace('.', '').find(str(part_number)) != -1:
#             result.append(f'Артикул {item[0]} в количестве '
#                       f'{str(item[-1]).replace(" ", "")} на '
#                       f'складе Глобал')
#     wb.close()
#     wb2.close()
#     return result


def open_xls_as_xlsx(filename):
    book = xlrd.open_workbook(filename)
    sheet = book.sheet_by_index(0)

    book1 = Workbook()
    sheet1 = book1.get_sheet_by_name(book1.get_sheet_names()[0])
    for row in range(sheet.nrows):
        sheet1.append(sheet.row_values(row))
    if filename == './downloads/Остатки.xls':
        sheet1.delete_rows(1, 2)
    if filename == './downloads/global.xls':
        sheet1.delete_rows(1, 4)
        sheet1.delete_cols(1, 2)
    return book1


def download_stocks():
    filename_eparh = URL_EPARH.split('/')[-1]
    downloads_dir = BASE_DIR / 'downloads'
    downloads_dir.mkdir(exist_ok=True)
    clear_folder(downloads_dir)
    archive_path_eparh = downloads_dir / filename_eparh
    archive_path_global = downloads_dir / 'global.xls'
    response = requests.get(URL_EPARH)
    with open(archive_path_eparh, 'wb') as file:
        file.write(response.content)
    with ZipFile(archive_path_eparh, "r") as myzip:
        myzip.extractall(path=downloads_dir)
    response = requests.get(URL_GLOBAL)
    with open(archive_path_global, 'wb') as file:
        file.write(response.content)
    open_xls_as_xlsx('./downloads/Остатки.xls').save(
        './downloads/eparh.xlsx')
    open_xls_as_xlsx('./downloads/global.xls').save(
        './downloads/global.xlsx')
