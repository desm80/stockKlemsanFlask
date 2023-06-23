import os
from zipfile import ZipFile

import requests
import xlrd
from openpyxl import Workbook, load_workbook

from klemsan_app import db, make_celery
from klemsan_app.models import KlemsanStock
from settings import URL_EPARH, BASE_DIR, URL_GLOBAL


celery = make_celery()


def clear_folder(folder):
    for f in os.listdir(folder):
        os.remove(os.path.join(folder, f))


def open_xls_as_xlsx(filename):
    book = xlrd.open_workbook(filename)
    sheet = book.sheet_by_index(0)

    book1 = Workbook()
    sheet1 = book1.get_sheet_by_name(book1.get_sheet_names()[0])
    for row in range(sheet.nrows):
        sheet1.append(sheet.row_values(row))
    if filename == './downloads/Остатки.xls':
        sheet1.delete_rows(1, 2)
    if filename == './downloads/global.xls':
        sheet1.delete_rows(1, 4)
        sheet1.delete_cols(1, 2)
    return book1


@celery.task(name='download_files')
def download_stocks():
    filename_eparh = URL_EPARH.split('/')[-1]
    downloads_dir = BASE_DIR / 'downloads'
    downloads_dir.mkdir(exist_ok=True)
    clear_folder(downloads_dir)
    archive_path_eparh = downloads_dir / filename_eparh
    archive_path_global = downloads_dir / 'global.xls'
    response = requests.get(URL_EPARH)
    with open(archive_path_eparh, 'wb') as file:
        file.write(response.content)
    with ZipFile(archive_path_eparh, "r") as myzip:
        myzip.extractall(path=downloads_dir)
    response = requests.get(URL_GLOBAL)
    with open(archive_path_global, 'wb') as file:
        file.write(response.content)
    open_xls_as_xlsx('./downloads/Остатки.xls').save(
        './downloads/eparh.xlsx')
    open_xls_as_xlsx('./downloads/global.xls').save(
        './downloads/global.xlsx')


@celery.task(name='files_to_base')
def xlsx_to_base(files):
    """
    Читает данные из файлов и создает БД по товарным остаткам.
    """
    objects = []
    for file in files:

        wb = load_workbook(file)
        sheet = wb[wb.sheetnames[0]]
        store = file.split('/')[-1].split('.')[0]
        for item in sheet.values:
            objects.append(
                KlemsanStock(
                    part_number=str(item[0]).replace('.', ''),
                    amount=item[-1],
                    store=store
                )
            )
        wb.close()
    db.session.add_all(objects)
    db.session.commit()
