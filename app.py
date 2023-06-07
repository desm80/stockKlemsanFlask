from http import HTTPStatus

from flask import Flask, render_template
from flask_migrate import Migrate
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook

from forms import DownloadForm, PartNumberForm
from settings import Config
from utils import download_stocks

app = Flask(__name__)
app.config.from_object(Config)
db = SQLAlchemy(app)
migrate = Migrate(app, db)


class KlemsanStock(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    part_number = db.Column(db.String(16), nullable=False)
    amount = db.Column(db.Integer)
    store = db.Column(db.String(16))


@app.route('/', methods=['GET', 'POST'])
def klemsan_view():
    """
    Рендеринг формы и обработка запроса на проверку наличия на складах.
    """
    form = PartNumberForm()
    if form.validate_on_submit():
        part_number = form.part_number.data
        result = KlemsanStock.query.filter(
            KlemsanStock.part_number.contains(part_number)).all()
        return render_template('klemsan.html', form=form, result=result)
    return render_template('klemsan.html', form=form)


@app.route('/download', methods=['GET', 'POST'])
def download_view():
    """
    Рендеринг формы с кнопкой для загрузки складских остатков и вывода
    сообщения, если загрузка прошла успешно.
    """
    form = DownloadForm()
    if form.validate_on_submit():
        download_stocks()
        db.drop_all()
        db.create_all()
        xlsx_to_base(['./downloads/eparh.xlsx',
                      './downloads/global.xlsx'])
        message = 'Остатки успешно загружены!'

        return render_template('download.html', form=form, message=message)
    return render_template('download.html', form=form)


@app.errorhandler(HTTPStatus.NOT_FOUND)
def page_not_found(error):
    """
    Кастомный обработчик ошибки 404.
    """
    return render_template('404.html'), HTTPStatus.NOT_FOUND


@app.errorhandler(HTTPStatus.INTERNAL_SERVER_ERROR)
def internal_error(error):
    """
    Кастомный обработчик ошибки 500.
    """
    return render_template('500.html'), HTTPStatus.INTERNAL_SERVER_ERROR


def xlsx_to_base(files):
    """
    Читает данные из файлов и создает БД по товарным остаткам.
    """
    objects = []
    for file in files:
        wb = load_workbook(file)
        sheet = wb[wb.sheetnames[0]]
        store = file.split('/')[-1].split('.')[0]
        for item in sheet.values:
            objects.append(
                KlemsanStock(
                    part_number=str(item[0]).replace('.', ''),
                    amount=item[-1],
                    store=store
                )
            )
        wb.close()
    db.session.add_all(objects)
    db.session.commit()


if __name__ == '__main__':
    app.run()
